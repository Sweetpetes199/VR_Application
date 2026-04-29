"""
QSS Device Intake — Crestron Setup Report Writer
Generates a formatted .xlsx file for a Crestron device setup run.

Workbook layout:
  Sheet "Summary"   — one row per device (device type, hostname, serial, overall, duration)
  Sheet "<Hostname>" — one row per step with Result, Reason, Notes, Duration

Usage:
    from crestron_report import write_crestron_excel
    path = write_crestron_excel(payload, output_dir)
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_RED    = PatternFill("solid", fgColor="FFC7CE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_BLUE   = PatternFill("solid", fgColor="BDD7EE")
_GREY   = PatternFill("solid", fgColor="D9D9D9")

_BOLD   = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

_SUMMARY_HEADERS = [
    "Timestamp", "Device Type", "Hostname", "Serial Number", "TSID",
    "Session ID", "Overall", "Duration (s)", "Failed Steps",
]

_STEP_HEADERS = [
    "Step", "Result", "Reason", "Notes", "Duration (s)",
]


def _fill_for_result(result: str) -> PatternFill:
    r = (result or "").upper()
    if r == "PASS": return _GREEN
    if r == "FAIL": return _RED
    if r == "SKIP": return _YELLOW
    return _GREY


def _header_row(ws, headers: list, row: int = 1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = _BOLD
        cell.fill      = _BLUE
        cell.alignment = _CENTER
        cell.border    = _THIN_BORDER


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)


def write_crestron_excel(payload: dict, output_dir: str) -> str:
    """
    Write (or append to) the Crestron setup Excel report.

    One file per calendar day: crestron_setup_YYYY-MM-DD.xlsx
    Summary sheet gets one new row per device.
    A dedicated sheet per device (named by hostname) lists all steps.

    Returns the absolute path of the written file.
    """
    from datetime import datetime

    ts          = payload.get("Timestamp", "")
    date_str    = ts[:10] if ts else datetime.now().strftime("%Y-%m-%d")
    filename    = f"crestron_setup_{date_str}.xlsx"
    path        = os.path.join(output_dir, filename)

    device_type  = payload.get("DeviceType",   "UNKNOWN")
    hostname     = payload.get("Hostname",      "UNKNOWN")
    serial       = payload.get("SerialNumber",  "")
    tsid         = payload.get("TSID",          "")
    session_id   = payload.get("SessionId",     "")
    overall      = payload.get("Overall",       "FAIL").upper()
    duration     = payload.get("DurationSec",   0)
    steps        = payload.get("Steps",         [])

    failed_steps = "; ".join(s["Name"] for s in steps if s.get("Result", "").upper() != "PASS")

    # ── Open or create workbook ───────────────────────────────────────────────
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Summary sheet ─────────────────────────────────────────────────────────
    if "Summary" not in wb.sheetnames:
        ws_sum = wb.create_sheet("Summary", 0)
        _header_row(ws_sum, _SUMMARY_HEADERS)
        ws_sum.freeze_panes = "A2"
    else:
        ws_sum = wb["Summary"]

    sum_row = ws_sum.max_row + 1
    values  = [ts, device_type, hostname, serial, tsid, session_id, overall, duration, failed_steps]
    fill    = _fill_for_result(overall)

    for col, val in enumerate(values, 1):
        cell = ws_sum.cell(row=sum_row, column=col, value=val)
        cell.fill      = fill
        cell.alignment = _LEFT
        cell.border    = _THIN_BORDER

    _auto_width(ws_sum)

    # ── Per-device sheet ──────────────────────────────────────────────────────
    # Sheet name: "hostname" — truncated to 31 chars (Excel limit), dupes get suffix
    sheet_name = hostname[:28]
    if sheet_name in wb.sheetnames:
        sheet_name = f"{sheet_name[:25]}_{sum_row}"

    ws_dev = wb.create_sheet(sheet_name)

    # Device info block (rows 1-6)
    info_pairs = [
        ("Device Type",   device_type),
        ("Hostname",      hostname),
        ("Serial Number", serial),
        ("TSID",          tsid),
        ("Session ID",    session_id),
        ("Overall",       overall),
        ("Duration (s)",  duration),
        ("Timestamp",     ts),
    ]
    for i, (label, val) in enumerate(info_pairs, 1):
        ws_dev.cell(row=i, column=1, value=label).font  = _BOLD
        cell = ws_dev.cell(row=i, column=2, value=val)
        cell.alignment = _LEFT
        if label == "Overall":
            cell.fill = _fill_for_result(val)

    blank_row = len(info_pairs) + 2

    # Step table header
    _header_row(ws_dev, _STEP_HEADERS, row=blank_row)
    ws_dev.freeze_panes = f"A{blank_row + 1}"

    # Step rows
    for s in steps:
        r    = ws_dev.max_row + 1
        fill = _fill_for_result(s.get("Result", ""))
        vals = [
            s.get("Name",        ""),
            s.get("Result",      ""),
            s.get("Reason",      ""),
            s.get("Notes",       ""),
            s.get("DurationSec", 0),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws_dev.cell(row=r, column=col, value=val)
            cell.fill      = fill
            cell.alignment = _LEFT
            cell.border    = _THIN_BORDER

    _auto_width(ws_dev)

    wb.save(path)
    return os.path.abspath(path)
