"""
Generates a formatted Excel workbook from a ScanSession payload.
One row per DeviceRecord — columns match the sticker fields exactly.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")  # dark navy
PASS_FILL    = PatternFill("solid", fgColor="C6EFCE")  # light green
FAIL_FILL    = PatternFill("solid", fgColor="FFC7CE")  # light red
PENDING_FILL = PatternFill("solid", fgColor="FFEB9C")  # light yellow
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUMNS = [
    ("Box ID",         14),
    ("Room Name",      18),
    ("Serial Number",  24),
    ("MAC Address",    20),
    ("Pass / Fail",    12),
    ("Sticker Applied",16),
    ("Scanned At",     22),
    ("Print Job ID",   14),
]


def write_session_excel(session: dict, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Device Intake"

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"QSS Device Intake — Session {session.get('SessionId', '')}  |  "
        f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    )
    title_cell.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    title_cell.fill      = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Column headers ───────────────────────────────────────────────────────
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 20

    # ── Data rows ────────────────────────────────────────────────────────────
    records = session.get("Records", [])
    for row_idx, record in enumerate(records, start=3):
        pass_fail = record.get("PassFail", "PENDING").upper()

        row_fill = (
            PASS_FILL    if pass_fail == "PASS"    else
            FAIL_FILL    if pass_fail == "FAIL"    else
            PENDING_FILL
        )

        row_data = [
            record.get("BoxId", ""),
            record.get("RoomName", ""),
            record.get("SerialNumber", ""),
            record.get("MacAddress", ""),
            pass_fail,
            "YES" if record.get("StickerApplied") else "NO",
            record.get("ScannedAt", ""),
            record.get("PrintJobId", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = row_fill if col_idx == 5 else WHITE_FILL
            cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left",
                                       vertical="center")
            cell.border    = THIN_BORDER

    # ── Summary row ──────────────────────────────────────────────────────────
    summary_row = len(records) + 3
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    summary_cell = ws[f"A{summary_row}"]
    total   = len(records)
    passed  = sum(1 for r in records if r.get("PassFail","").upper() == "PASS")
    failed  = sum(1 for r in records if r.get("PassFail","").upper() == "FAIL")
    pending = total - passed - failed
    summary_cell.value = (
        f"Total: {total}   |   PASS: {passed}   FAIL: {failed}   PENDING: {pending}"
    )
    summary_cell.font      = Font(name="Calibri", bold=True, size=10)
    summary_cell.alignment = Alignment(horizontal="center")

    # Freeze panes below header
    ws.freeze_panes = "A3"

    wb.save(output_path)
    print(f"[Excel] Saved → {output_path}  ({total} records)")
